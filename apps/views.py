from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.models import User, Group
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count, Q  # se for usar no gráfico

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from datetime import datetime
import pandas as pd          # <--- NOVA LINHA
from io import BytesIO

from .models import (
    Evento, Documento, Recado, DocumentoPrivado, EventoPrivado, RecadoInterno,
    PeriodoAula, RegistroFalta, Video, Turma, Aluno, RegistroFaltaAluno,
    RegistroOcorrenciaAluno, LogLogin, Professor  # ← ADICIONE APENAS Professor
)

from .forms import (
    RecadoInternoForm, DocumentoPrivadoForm, EventoPrivadoForm,
    RegistroOcorrenciaForm
)

# ===== NOVO IMPORT DO WHATSAPP =====
from apps.utils import enviar_whatsapp

# Função de teste para verificar se o usuário está no grupo "Equipe Diretiva"
def pertence_ao_grupo_equipe_diretiva(user):
    if user.is_authenticated:
        return user.groups.filter(name__iexact='Equipe Diretiva').exists()
    return False

# Função de teste para verificar se o usuário está no grupo "Equipe Diretiva"
def pertence_ao_grupo_equipe_diretiva(user):
    if user.is_authenticated:
        return user.groups.filter(name__iexact='Equipe Diretiva').exists()
    return False

def home(request):
    eventos = Evento.objects.all().order_by('data')[:5]
    recados = Recado.objects.filter(fixado=True)
    documentos = Documento.objects.all().order_by('-criado_em')  # ← ALTERADO
    videos = Video.objects.all().order_by('-criado_em')[:3]

    return render(request, 'home.html', {
        'eventos': eventos,
        'recados': recados,
        'documentos': documentos,
        'videos': videos,
        'user': request.user,
    })

from django.db.models import Count
from calendar import monthrange
from datetime import datetime
from .models import Aluno, RegistroOcorrenciaAluno, RegistroFaltaAluno  # se ainda não estiver importado
from .forms import RelatorioFaltasForm
from .forms import RelatorioFaltasForm, RegistroOcorrenciaForm

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def painel_equipe(request):
    # --- Conteúdos existentes ---
    documentos_privados = DocumentoPrivado.objects.all().order_by('-criado_em')
    recados_internos = RecadoInterno.objects.all().order_by('-criado_em')[:20]
    eventos_privados = EventoPrivado.objects.all().order_by('data_inicio')
    turmas = Turma.objects.all().order_by('nome')

    # ===== GRÁFICO DE FALTAS POR TURNO =====
    mes = int(request.GET.get('mes', datetime.now().month))
    ano = int(request.GET.get('ano', datetime.now().year))

    # DIAGNÓSTICO - LOG 1
    print(f"=== GRÁFICO - Mês/Ano: {mes}/{ano} ===")

    # Total de alunos por turno
    total_manha = Aluno.objects.filter(turma__turno='manha').count()
    total_tarde = Aluno.objects.filter(turma__turno='tarde').count()

    # Evitar divisão por zero
    if total_manha == 0:
        total_manha = 1
    if total_tarde == 0:
        total_tarde = 1

    print(f"Total alunos manhã: {total_manha}, tarde: {total_tarde}")

    # Número de dias no mês
    _, num_dias = monthrange(ano, mes)

    # Inicializar contadores por dia
    faltas_manha_por_dia = {d: 0 for d in range(1, num_dias + 1)}
    faltas_tarde_por_dia = {d: 0 for d in range(1, num_dias + 1)}

    # Consulta faltas (RegistroOcorrenciaAluno) com agrupamento por dia e turno
    ocorrencias = RegistroOcorrenciaAluno.objects.filter(
        faltou=True,
        data__year=ano,
        data__month=mes
    ).select_related('aluno__turma')

    # DIAGNÓSTICO - LOG 2
    total_faltas = ocorrencias.count()
    print(f"Total de faltas encontradas: {total_faltas}")

    for occ in ocorrencias:
        dia = occ.data.day
        # AGORA USA O CAMPO turno DA PRÓPRIA OCORRÊNCIA
        if occ.turno == 'manha':
            faltas_manha_por_dia[dia] += 1
        elif occ.turno == 'tarde':
            faltas_tarde_por_dia[dia] += 1

    # DIAGNÓSTICO - LOG 3
    total_manha_contado = sum(faltas_manha_por_dia.values())
    total_tarde_contado = sum(faltas_tarde_por_dia.values())
    print(f"Faltas contadas - manhã: {total_manha_contado}, tarde: {total_tarde_contado}")

    # Calcula percentuais para cada dia
    percentuais_manha = []
    percentuais_tarde = []
    for dia in range(1, num_dias + 1):
        perc_manha = (faltas_manha_por_dia[dia] / total_manha) * 100
        perc_tarde = (faltas_tarde_por_dia[dia] / total_tarde) * 100
        percentuais_manha.append(round(perc_manha, 2))
        percentuais_tarde.append(round(perc_tarde, 2))

    # DIAGNÓSTICO - LOG 4
    print(f"Primeiros 5 percentuais manhã: {percentuais_manha[:5]}")
    print(f"Primeiros 5 percentuais tarde: {percentuais_tarde[:5]}")
    print(f"Últimos 5 percentuais manhã: {percentuais_manha[-5:]}")
    print(f"Últimos 5 percentuais tarde: {percentuais_tarde[-5:]}")

    dias = list(range(1, num_dias + 1))
    # ===== FIM DO GRÁFICO =====

    # ===== RANKING DE FALTAS POR TURMA =====
    # Captura mês e ano do ranking (GET) ou usa o mês/ano atual
    mes_ranking = int(request.GET.get('mes_ranking', datetime.now().month))
    ano_ranking = int(request.GET.get('ano_ranking', datetime.now().year))

    # Busca faltas no período
    faltas_ranking = RegistroOcorrenciaAluno.objects.filter(
        faltou=True,
        data__year=ano_ranking,
        data__month=mes_ranking
    ).select_related('aluno', 'aluno__turma')

    # Conta faltas por turma e separa por turno
    turmas_manha = {}
    turmas_tarde = {}

    for falta in faltas_ranking:
        turma_nome = falta.aluno.turma.nome
        turno = falta.aluno.turma.turno
        if turno == 'manha':
            turmas_manha[turma_nome] = turmas_manha.get(turma_nome, 0) + 1
        elif turno == 'tarde':
            turmas_tarde[turma_nome] = turmas_tarde.get(turma_nome, 0) + 1

    # Ordena por total de faltas (decrescente)
    ranking_manha = sorted(turmas_manha.items(), key=lambda x: x[1], reverse=True)
    ranking_tarde = sorted(turmas_tarde.items(), key=lambda x: x[1], reverse=True)

    # Separa dados para os gráficos
    turmas_manha_lista = [item[0] for item in ranking_manha]
    totais_manha_lista = [item[1] for item in ranking_manha]
    turmas_tarde_lista = [item[0] for item in ranking_tarde]
    totais_tarde_lista = [item[1] for item in ranking_tarde]
    # ===== FIM DO RANKING =====

    # ===== CONTEXTO =====
    context = {
        'usuario': request.user,
        'documentos_privados': documentos_privados,
        'recados_internos': recados_internos,
        'eventos_privados': eventos_privados,
        'turmas': turmas,
        'mes_atual': datetime.now().month,
        'ano_atual': datetime.now().year,
        # Dados para o gráfico de faltas diárias
        'dias': dias,
        'percentuais_manha': percentuais_manha,
        'percentuais_tarde': percentuais_tarde,
        'total_manha': total_manha,
        'total_tarde': total_tarde,
        'mes': mes,
        'ano': ano,
        'meses': range(1, 13),
        'anos': range(2024, 2027),
        # Dados para o ranking
        'ranking_manha': ranking_manha,
        'ranking_tarde': ranking_tarde,
        'turmas_manha': turmas_manha_lista,
        'totais_manha': totais_manha_lista,
        'turmas_tarde': turmas_tarde_lista,
        'totais_tarde': totais_tarde_lista,
        'mes_ranking': mes_ranking,
        'ano_ranking': ano_ranking,
        'total_faltas_periodo': sum(totais_manha_lista) + sum(totais_tarde_lista),
    }
    return render(request, 'painel_equipe.html', context)

# ===== NOVA VIEW: RELATÓRIO DE FALTAS POR ALUNO =====
@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def relatorio_faltas(request):
    form = RelatorioFaltasForm()
    erro = None

    if request.method == 'POST':
        form = RelatorioFaltasForm(request.POST)
        if form.is_valid():
            turma = form.cleaned_data['turma']
            identificador = form.cleaned_data['identificador']

            # Buscar aluno
            aluno = None
            if identificador.isdigit():
                aluno = Aluno.objects.filter(turma=turma, numero=int(identificador)).first()
            else:
                aluno = Aluno.objects.filter(turma=turma, nome__icontains=identificador).first()

            if not aluno:
                erro = 'Aluno não encontrado nesta turma.'
            else:
                # Busca as ocorrências onde o aluno faltou (faltou=True)
                faltas = RegistroOcorrenciaAluno.objects.filter(
                    aluno=aluno,
                    faltou=True
                ).order_by('-data')
                total_faltas = faltas.count()  # cada ocorrência conta como 1 falta
                return render(request, 'relatorio_faltas.html', {
                    'aluno': aluno,
                    'faltas': faltas,
                    'total_faltas': total_faltas
                })

    return render(request, 'form_relatorio_faltas.html', {'form': form, 'erro': erro})

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def upload_documento_privado(request):
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        categoria = request.POST.get('categoria')
        descricao = request.POST.get('descricao')
        arquivo = request.FILES.get('arquivo')
        if titulo and arquivo:
            DocumentoPrivado.objects.create(
                titulo=titulo,
                arquivo=arquivo,
                categoria=categoria,
                descricao=descricao,
                criado_por=request.user
            )
    return redirect('painel_equipe')



@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def criar_evento_privado(request):
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        descricao = request.POST.get('descricao')
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        local = request.POST.get('local')
        if titulo and data_inicio:
            EventoPrivado.objects.create(
                titulo=titulo,
                descricao=descricao,
                data_inicio=data_inicio,
                data_fim=data_fim if data_fim else None,
                local=local,
                criado_por=request.user
            )
    return redirect('painel_equipe')

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def enviar_recado_interno(request):
    if request.method == 'POST':
        mensagem = request.POST.get('mensagem')
        arquivo = request.FILES.get('arquivo')
        if mensagem:
            RecadoInterno.objects.create(
                mensagem=mensagem,
                arquivo=arquivo,
                criado_por=request.user
            )
    return redirect('painel_equipe')

# ===== NOVAS VIEWS PARA CONTROLE DE FALTAS =====

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def controle_faltas(request):
    professores = User.objects.filter(groups__name__iexact='Equipe Diretiva').exclude(id=request.user.id).order_by('first_name')

    # Captura e converte os valores da URL
    mes_str = request.GET.get('mes')
    ano_str = request.GET.get('ano')

    if mes_str:
        mes = int(mes_str)
    else:
        mes = timezone.now().month

    if ano_str:
        ano = int(ano_str)
    else:
        ano = timezone.now().year

    print(f"🔍 Mês recebido: {mes}, Ano recebido: {ano}")
    print(f"🔍 URL completa: {request.get_full_path()}")

    faltas = RegistroFalta.objects.filter(data__month=mes, data__year=ano).select_related('professor').order_by('-data')
    periodos = PeriodoAula.objects.all().order_by('ordem')
    resumo = []
    for professor in professores:
        faltas_prof = [f for f in faltas if f.professor.id == professor.id]
        total_minutos = sum(f.minutos_faltados() for f in faltas_prof)
        total_faltas = len([f for f in faltas_prof if f.tipo == 'falta'])
        total_atrasos = len([f for f in faltas_prof if f.tipo == 'atraso'])
        resumo.append({
            'nome': professor.get_full_name() or professor.username,
            'total_minutos': total_minutos,
            'total_horas': round(total_minutos / 60, 1),
            'total_faltas': total_faltas,
            'total_atrasos': total_atrasos,
        })
    resumo = sorted(resumo, key=lambda x: x['nome'])
    context = {
        'professores': professores,
        'faltas': faltas,
        'periodos': periodos,
        'resumo': resumo,
        'mes': mes,
        'ano': ano,
        'meses': range(1, 13),
        'anos': [2026, 2027],
    }
    return render(request, 'controle_faltas.html', context)

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def registrar_falta(request):
    professores = User.objects.filter(groups__name__iexact='Professores')
    if request.method == 'GET':
        return render(request, 'registrar_falta.html', {'professores': professores})
    if request.method == 'POST':
        professor_id = request.POST.get('professor')
        data = request.POST.get('data')
        horario_previsto = request.POST.get('horario_previsto')
        horario_real = request.POST.get('horario_real') or None
        tipo = request.POST.get('tipo')
        observacao = request.POST.get('observacao', '')
        try:
            professor = User.objects.get(id=professor_id)
            data_obj = datetime.strptime(data, '%Y-%m-%d').date()
            previsto_obj = datetime.strptime(horario_previsto, '%H:%M').time()
            real_obj = None
            if horario_real:
                real_obj = datetime.strptime(horario_real, '%H:%M').time()
            minutos_atraso = 0
            if real_obj and tipo == 'atraso':
                previsto_min = previsto_obj.hour * 60 + previsto_obj.minute
                real_min = real_obj.hour * 60 + real_obj.minute
                minutos_atraso = max(0, real_min - previsto_min)
            RegistroFalta.objects.create(
                professor=professor,
                data=data_obj,
                horario_previsto=previsto_obj,
                horario_real=real_obj,
                tipo=tipo,
                minutos_atraso=minutos_atraso,
                observacao=observacao,
                registrado_por=request.user
            )
            messages.success(request, 'Registro salvo com sucesso!')
            return redirect('controle_faltas')
        except User.DoesNotExist:
            messages.error(request, 'Professor não encontrado!')
        except Exception as e:
            messages.error(request, f'Erro ao salvar: {str(e)}')
        return redirect('controle_faltas')

# =============================================================================
# CONTROLE DE FALTAS - ALUNOS (LISTAGEM)
# =============================================================================
from django.utils import timezone

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def controle_faltas_alunos(request):
    # Filtros
    mes = request.GET.get('mes', timezone.now().month)
    ano = request.GET.get('ano', timezone.now().year)

    # Busca ocorrências do período
    ocorrencias = RegistroOcorrenciaAluno.objects.filter(
        data__month=mes,
        data__year=ano
    ).select_related('aluno', 'aluno__turma').order_by('-data', '-horario_chegada')

    context = {
        'ocorrencias': ocorrencias,
        'mes': mes,
        'ano': ano,
        'meses': range(1, 13),
        'anos': [2026, 2027, 2028],
    }
    return render(request, 'controle_faltas_alunos.html', context)

# =============================================================================
# EDITAR OCORRÊNCIA DE ALUNO
# =============================================================================

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def editar_ocorrencia_aluno(request, pk):
    from datetime import datetime
    ocorrencia = get_object_or_404(RegistroOcorrenciaAluno, id=pk)

    if request.method == 'POST':
        # 🔧 1. SALVA OS CAMPOS DA BUSCA ATIVA MESMO ANTES DA VALIDAÇÃO
        ocorrencia.responsavel_contatado = request.POST.get('responsavel_contatado', '')
        ocorrencia.alegado_responsavel = request.POST.get('alegado_responsavel', '')
        horario_contato = request.POST.get('horario_contato', '')
        if horario_contato:
            try:
                ocorrencia.horario_contato = datetime.strptime(horario_contato, '%H:%M').time()
            except:
                ocorrencia.horario_contato = None
        else:
            ocorrencia.horario_contato = None
        ocorrencia.save(update_fields=['responsavel_contatado', 'alegado_responsavel', 'horario_contato'])

        # 2. Agora processa o formulário normalmente (se quiser salvar outras alterações)
        form = RegistroOcorrenciaForm(request.POST, instance=ocorrencia)
        if form.is_valid():
            # Salva as outras alterações (turma, número, data, etc.)
            form.save()
            messages.success(request, 'Ocorrência atualizada com sucesso!')
        else:
            # Mesmo com erro, os dados da Busca Ativa já foram salvos
            messages.warning(request, 'Busca Ativa salva, mas outros dados apresentaram erro. Verifique o formulário.')

        # Redireciona de volta para a Busca Ativa mantendo os filtros
        mes = request.GET.get('mes', '')
        ano = request.GET.get('ano', '')
        turma_id = request.GET.get('turma', '')
        url = '/busca-ativa/'
        params = []
        if mes:
            params.append(f'mes={mes}')
        if ano:
            params.append(f'ano={ano}')
        if turma_id:
            params.append(f'turma={turma_id}')
        if params:
            url += '?' + '&'.join(params)
        return redirect(url)
    else:
        # GET: exibe o formulário normalmente
        form = RegistroOcorrenciaForm(instance=ocorrencia)
        form.fields['turma'].initial = ocorrencia.aluno.turma
        form.fields['numero_aluno'].initial = ocorrencia.aluno.numero
        form.fields['nome_aluno'].initial = ocorrencia.aluno.nome
        form.fields['data'].initial = ocorrencia.data

    return render(request, 'editar_ocorrencia_aluno.html', {
        'form': form,
        'ocorrencia': ocorrencia,
    })

# =============================================================================
# EXCLUIR OCORRÊNCIA DE ALUNO
# =============================================================================
@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def excluir_ocorrencia_aluno(request, pk):
    ocorrencia = get_object_or_404(RegistroOcorrenciaAluno, id=pk)
    if request.method == 'POST':
        ocorrencia.delete()
        messages.success(request, 'Ocorrência excluída com sucesso!')
        return redirect('controle_faltas_alunos')
    return render(request, 'confirmar_exclusao_ocorrencia.html', {'ocorrencia': ocorrencia})


# =============================================================================
# VIEW PARA EDITAR UM REGISTRO DE FALTA
# =============================================================================
@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def editar_falta(request, falta_id):
    falta = get_object_or_404(RegistroFalta, id=falta_id)
    professores = User.objects.filter(groups__name__iexact='Professores')

    if request.method == 'POST':
        professor_id = request.POST.get('professor')
        data = request.POST.get('data')
        horario_previsto = request.POST.get('horario_previsto')
        horario_real = request.POST.get('horario_real') or None
        tipo = request.POST.get('tipo')
        observacao = request.POST.get('observacao', '')

        try:
            professor = User.objects.get(id=professor_id)
            falta.professor = professor
            falta.data = datetime.strptime(data, '%Y-%m-%d').date()
            falta.horario_previsto = datetime.strptime(horario_previsto, '%H:%M').time()
            falta.horario_real = datetime.strptime(horario_real, '%H:%M').time() if horario_real else None
            falta.tipo = tipo
            falta.observacao = observacao

            # Recalcular minutos de atraso
            if falta.horario_real and tipo == 'atraso':
                previsto_min = falta.horario_previsto.hour * 60 + falta.horario_previsto.minute
                real_min = falta.horario_real.hour * 60 + falta.horario_real.minute
                falta.minutos_atraso = max(0, real_min - previsto_min)
            else:
                falta.minutos_atraso = 0

            falta.save()
            messages.success(request, 'Registro atualizado com sucesso!')
            return redirect('controle_faltas')
        except Exception as e:
            messages.error(request, f'Erro ao atualizar: {str(e)}')
            return redirect('editar_falta', falta_id=falta.id)

    # GET - exibe o formulário
    context = {
        'falta': falta,
        'professores': professores,
        'hoje': falta.data.strftime('%Y-%m-%d'),
        'horario_previsto': falta.horario_previsto.strftime('%H:%M') if falta.horario_previsto else '',
        'horario_real': falta.horario_real.strftime('%H:%M') if falta.horario_real else '',
    }
    return render(request, 'editar_falta.html', context)
# =============================================================================
# VIEWS PARA EXCLUSÃO DE REGISTROS DE FALTA
# =============================================================================

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def excluir_falta(request, falta_id):
    """Exclui um único registro de falta."""
    falta = get_object_or_404(RegistroFalta, id=falta_id)
    falta.delete()
    messages.success(request, 'Registro excluído com sucesso.')
    return redirect('controle_faltas')


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def excluir_faltas_selecionadas(request):
    """Exclui múltiplos registros de falta enviados por POST."""
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            RegistroFalta.objects.filter(id__in=ids).delete()
            messages.success(request, f'{len(ids)} registro(s) excluído(s) com sucesso.')
        else:
            messages.warning(request, 'Nenhum registro selecionado para exclusão.')
    return redirect('controle_faltas')

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def relatorio_faltas_por_aluno(request):
    # Pega mês e ano da URL (ou atual)
    mes = int(request.GET.get('mes', timezone.now().month))
    ano = int(request.GET.get('ano', timezone.now().year))

    # Filtra ocorrências com falta no período
    faltas = RegistroOcorrenciaAluno.objects.filter(
        faltou=True,
        data__year=ano,
        data__month=mes
    ).select_related('aluno', 'aluno__turma')

    # Agrupa por aluno e conta faltas
    alunos_faltas = {}
    for falta in faltas:
        aluno = falta.aluno
        key = (aluno.turma.nome, aluno.numero, aluno.nome)
        alunos_faltas[key] = alunos_faltas.get(key, 0) + 1

    # Converte para lista ordenada
    dados = []
    for (turma, numero, nome), total in sorted(alunos_faltas.items(), key=lambda x: (x[0][0], x[0][1])):
        dados.append({
            'turma': turma,
            'numero': numero,
            'nome': nome,
            'faltas': total,
        })

    context = {
        'dados': dados,
        'mes': mes,
        'ano': ano,
        'meses': range(1, 13),
        'anos': range(2024, 2028),
    }
    return render(request, 'relatorio_faltas_por_aluno.html', context)


def exportar_relatorio_faltas(request):
    mes = int(request.GET.get('mes', timezone.now().month))
    ano = int(request.GET.get('ano', timezone.now().year))

    # Busca TODOS os registros (não só faltas)
    ocorrencias = RegistroOcorrenciaAluno.objects.filter(
        data__year=ano,
        data__month=mes
    ).select_related('aluno', 'aluno__turma').order_by('data', 'aluno__turma__nome', 'aluno__numero')

    # Prepara os dados para o Excel (COM ORDEM FIXA)
    data = []
    for ocorrencia in ocorrencias:
        data.append({
            'Data': ocorrencia.data.strftime('%d/%m/%Y'),
            'Turma': ocorrencia.aluno.turma.nome,
            'Nº': ocorrencia.aluno.numero,
            'Aluno': ocorrencia.aluno.nome,
            'Tipo': ocorrencia.get_tipo_ocorrencia_display(),
            'Horário Chegada': ocorrencia.horario_chegada.strftime('%H:%M') if ocorrencia.horario_chegada else '',
            'Atendido por': ocorrencia.atendido_por,
            'Motivo (aluno)': ocorrencia.motivo_alegado,
            'Responsável contatado': ocorrencia.responsavel_contatado,
            'Hora contato': ocorrencia.horario_contato.strftime('%H:%M') if ocorrencia.horario_contato else '',
            'Alegado (responsável)': ocorrencia.alegado_responsavel,
        })

    # Cria o DataFrame com a ORDEm das colunas DEFINIDA
    colunas_ordem = ['Data', 'Turma', 'Nº', 'Aluno', 'Tipo', 'Horário Chegada',
                     'Atendido por', 'Motivo (aluno)', 'Responsável contatado',
                     'Hora contato', 'Alegado (responsável)']

    df = pd.DataFrame(data, columns=colunas_ordem)

    # Se não houver dados, cria um DataFrame vazio com as colunas
    if df.empty:
        df = pd.DataFrame(columns=colunas_ordem)

    # Gera o Excel
    output = BytesIO()
    sheet_name = f'Ocorrencias_{mes:02d}_{ano}'
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        # Ajusta largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="relatorio_ocorrencias_{mes:02d}_{ano}.xlsx"'
    return response

# ===== FUNÇÕES PARA DIGITADORES =====
def grupo_digitadores(user):
    """Verifica se o usuário pertence ao grupo Digitadores"""
    return user.groups.filter(name='Digitadores').exists()

@login_required
@user_passes_test(grupo_digitadores, login_url='/')
def formulario_digitador(request):
    """View exclusiva para digitadores (apenas o formulário de ocorrências)"""
    # Recupera última data da sessão
    ultima_data_str = request.session.get('ultima_data_ocorrencia')
    if ultima_data_str:
        try:
            ultima_data = datetime.strptime(ultima_data_str, '%Y-%m-%d').date()
        except ValueError:
            ultima_data = timezone.now().date()
    else:
        ultima_data = timezone.now().date()

    # Recupera última turma da sessão
    ultima_turma_id = request.session.get('ultima_turma_ocorrencia_id')
    ultima_turma = None
    if ultima_turma_id:
        try:
            ultima_turma = Turma.objects.get(id=ultima_turma_id)
        except Turma.DoesNotExist:
            ultima_turma = None

    if request.method == 'POST':
        form = RegistroOcorrenciaForm(request.POST)
        if form.is_valid():
            turma = form.cleaned_data['turma']
            numero = form.cleaned_data['numero_aluno']

            aluno = Aluno.objects.filter(turma=turma, numero=numero).first()
            if not aluno:
                messages.error(request, f'Aluno número {numero} não encontrado na turma {turma.nome}!')
                return render(request, 'formulario_digitador.html', {
                    'form': form,
                    'ultimas_ocorrencias': RegistroOcorrenciaAluno.objects.select_related('aluno', 'aluno__turma').order_by('-data', '-horario_chegada')[:10]
                })

            ocorrencia = form.save(commit=False)

            # 🔧 CORREÇÃO 1: Pega o tipo de ocorrência do formulário
            tipo_ocorrencia = form.cleaned_data.get('tipo_ocorrencia')
            if tipo_ocorrencia == 'falta':
                ocorrencia.faltou = True
            else:
                ocorrencia.faltou = False

            # 🔧 CORREÇÃO 2: Pega o turno do formulário
            ocorrencia.turno = form.cleaned_data.get('turno', 'manha')

            if ocorrencia.horario_chegada == '':
                ocorrencia.horario_chegada = None
            if ocorrencia.horario_contato == '':
                ocorrencia.horario_contato = None

            ocorrencia.aluno = aluno
            ocorrencia.registrado_por = request.user
            ocorrencia.save()

            request.session['ultima_data_ocorrencia'] = ocorrencia.data.isoformat()
            request.session['ultima_turma_ocorrencia_id'] = turma.id
            request.session['ultima_turma_ocorrencia_nome'] = turma.nome

            messages.success(request, f'Ocorrência registrada para {aluno.nome} (Turma {turma.nome}, Nº {numero})')
            return redirect('formulario_digitador')
        else:
            messages.error(request, 'Erro no formulário. Verifique os dados.')
    else:
        initial_data = {
            'data': ultima_data.isoformat(),
            'faltou': True,
        }
        form = RegistroOcorrenciaForm(initial=initial_data)
        if ultima_turma:
            form.fields['turma'].initial = ultima_turma

    ultimas_ocorrencias = RegistroOcorrenciaAluno.objects.select_related('aluno', 'aluno__turma').order_by('-data', '-horario_chegada')[:10]

    return render(request, 'formulario_digitador.html', {
        'form': form,
        'ultimas_ocorrencias': ultimas_ocorrencias
    })
# ===== REDIRECIONAMENTO PERSONALIZADO PARA LOGIN =====
from django.contrib.auth.views import LoginView

class CustomLoginView(LoginView):
    def get_success_url(self):
        user = self.request.user

        # Registra o acesso no log
        from .models import LogLogin
        LogLogin.objects.create(
            usuario=user,
            ip=self.request.META.get('REMOTE_ADDR'),
            user_agent=self.request.META.get('HTTP_USER_AGENT', '')
        )

        if user.groups.filter(name='Digitadores').exists():
            return '/ocorrencias/registrar/'
        return '/painel-equipe/'


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def relatorio_faltas_mensal(request):
    mes = request.GET.get('mes', timezone.now().month)
    ano = request.GET.get('ano', timezone.now().year)
    professores = User.objects.filter(groups__name__iexact='Equipe Diretiva')
    relatorio = []
    for professor in professores:
        faltas = RegistroFalta.objects.filter(professor=professor, data__month=mes, data__year=ano).order_by('data', 'periodo__ordem')
        total_minutos = sum(f.minutos_faltados() for f in faltas)
        total_faltas = len([f for f in faltas if f.tipo == 'falta'])
        total_atrasos = len([f for f in faltas if f.tipo == 'atraso'])
        relatorio.append({
            'professor': professor.get_full_name() or professor.username,
            'faltas': faltas,
            'total_minutos': total_minutos,
            'total_horas': round(total_minutos / 60, 1),
            'total_faltas': total_faltas,
            'total_atrasos': total_atrasos,
        })
    context = {
        'relatorio': relatorio,
        'mes': mes,
        'ano': ano,
        'meses': range(1, 13),
        'anos': [2026, 2027],
    }
    return render(request, 'relatorio_faltas.html', context)

# ===== HISTÓRICO DE ACESSOS =====
@login_required
def historico_acessos(request):
    # Apenas superusuários podem acessar
    if not request.user.is_superuser:
        return redirect('painel_equipe')

    logs = LogLogin.objects.all().order_by('-data_hora')
    return render(request, 'historico_acessos.html', {'logs': logs})

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def lancar_ponto(request):
    from .models import PeriodoAula, RegistroPonto
    professores = User.objects.filter(groups__name__iexact='Equipe Diretiva').order_by('first_name')
    periodos = PeriodoAula.objects.all().order_by('ordem')
    data_atual = timezone.now().date()
    if request.method == 'POST':
        professor_id = request.POST.get('professor')
        data = request.POST.get('data')
        periodo_id = request.POST.get('periodo')
        acao = request.POST.get('acao')
        professor = User.objects.get(id=professor_id)
        periodo = PeriodoAula.objects.get(id=periodo_id)
        registro, created = RegistroPonto.objects.get_or_create(
            professor=professor,
            data=data,
            periodo=periodo,
            defaults={
                'horario_previsto_inicio': periodo.inicio,
                'horario_previsto_fim': periodo.fim,
                'registrado_por': request.user,
            }
        )
        if acao == 'entrada':
            registro.entrada_real = timezone.now()
            messages.success(request, f'Entrada registrada para {professor.get_full_name()} às {timezone.now().strftime("%H:%M")}')
        elif acao == 'saida':
            registro.saida_real = timezone.now()
            messages.success(request, f'Saída registrada para {professor.get_full_name()} às {timezone.now().strftime("%H:%M")}')
        registro.save()
        return redirect('lancar_ponto')
    context = {
        'professores': professores,
        'periodos': periodos,
        'data_atual': data_atual,
        'hoje': data_atual,
    }
    return render(request, 'lancar_ponto.html', context)

def detalhe_evento(request, evento_id):
    from .models import Evento
    evento = Evento.objects.get(id=evento_id)
    return render(request, 'detalhe_evento.html', {'evento': evento})

def detalhe_recado(request, recado_id):
    from .models import Recado
    recado = Recado.objects.get(id=recado_id)
    return render(request, 'detalhe_recado.html', {'recado': recado})

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def relatorio_ponto(request):
    from .models import RegistroPonto
    mes = request.GET.get('mes', timezone.now().month)
    ano = request.GET.get('ano', timezone.now().year)
    professores = User.objects.filter(groups__name__iexact='Equipe Diretiva')
    registros = RegistroPonto.objects.filter(data__month=mes, data__year=ano).select_related('professor', 'periodo').order_by('data', 'periodo_ordem')
    resumo = []
    for professor in professores:
        registros_prof = [r for r in registros if r.professor.id == professor.id]
        total_horas_trabalhadas = sum(r.horas_trabalhadas() for r in registros_prof)
        total_horas_previstas = sum(r.horas_previstas() for r in registros_prof)
        saldo_total = sum(r.saldo() for r in registros_prof)
        resumo.append({
            'nome': professor.get_full_name() or professor.username,
            'total_horas_trabalhadas': round(total_horas_trabalhadas, 2),
            'total_horas_previstas': round(total_horas_previstas, 2),
            'saldo': round(saldo_total, 2),
            'registros': registros_prof,
        })
    context = {
        'resumo': resumo,
        'registros': registros,
        'mes': mes,
        'ano': ano,
        'meses': range(1, 13),
        'anos': [2026, 2027, 2028, 2029, 2030],
    }
    return render(request, 'relatorio_ponto.html', context)

# =============================================================================
# FUNÇÃO PARA ARQUIVAR MÊS
# =============================================================================
@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def arquivar_mes(request):
    print("="*50)
    print("🚀 FUNÇÃO ARQUIVAR_MES FOI CHAMADA!")
    print("="*50)

    faltas = RegistroFalta.objects.all()
    print(f"📊 Total de registros encontrados: {faltas.count()}")

    if not faltas.exists():
        print("⚠️ Nenhum registro encontrado. Abortando missão.")
        messages.warning(request, "Nenhum registro para arquivar este mês.")
        return redirect('controle_faltas')

    hoje = datetime.now()
    mes_atual = hoje.strftime("%B").capitalize()
    ano_atual = hoje.strftime("%Y")
    nome_arquivo = f"faltas_{mes_atual}_{ano_atual}.xlsx"
    print(f"📁 Nome do arquivo a ser criado: {nome_arquivo}")

    caminho_pasta = '/home/gestaoluizzanchim/colegio/media/arquivos_mensais/'
    caminho_completo = os.path.join(caminho_pasta, nome_arquivo)
    print(f"📂 Caminho completo: {caminho_completo}")

    try:
        print("📄 Tentando criar o arquivo Excel com openpyxl...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{mes_atual} {ano_atual}"

        headers = ['Professor', 'Data', 'Dia', 'Previsto', 'Real', 'Status', 'Minutos', 'Observação']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        for row_num, falta in enumerate(faltas, 2):
            ws.cell(row=row_num, column=1, value=falta.professor.get_full_name() or falta.professor.username)
            ws.cell(row=row_num, column=2, value=falta.data.strftime('%d/%m/%Y'))
            ws.cell(row=row_num, column=3, value=falta.dia_semana)
            ws.cell(row=row_num, column=4, value=falta.horario_previsto.strftime('%H:%M') if falta.horario_previsto else '-')
            ws.cell(row=row_num, column=5, value=falta.horario_real.strftime('%H:%M') if falta.horario_real else '-')
            ws.cell(row=row_num, column=6, value=falta.get_tipo_display())
            ws.cell(row=row_num, column=7, value=f"{falta.minutos_faltados()} min")
            ws.cell(row=row_num, column=8, value=falta.observacao or '-')

        print("✅ Dados inseridos no Excel com sucesso.")
        print("💾 Tentando salvar o arquivo...")
        wb.save(caminho_completo)
        print("✅ Arquivo salvo com sucesso no disco!")

    except Exception as e:
        print(f"❌❌❌ ERRO NA CRIAÇÃO DO EXCEL: {e}")
        messages.error(request, f"Erro ao criar arquivo Excel: {e}")
        return redirect('controle_faltas')

    messages.success(request, f"Mês de {mes_atual} arquivado com sucesso!")
    print("="*50)
    print("🎉🎉🎉 FUNÇÃO ARQUIVAR_MES CONCLUÍDA COM SUCESSO! 🎉🎉🎉")
    print("="*50)
    return redirect('controle_faltas')

# =============================================================================
# FUNÇÃO PARA EXPORTAR EXCEL (SEM APAGAR NADA)
# =============================================================================
@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def exportar_excel_faltas(request):
    """Gera um arquivo Excel com os registros ATUAIS (sem apagar nada)"""

    # Pega os registros do mês atual (mesmo filtro da página)
    mes = request.GET.get('mes', timezone.now().month)
    ano = request.GET.get('ano', timezone.now().year)
    faltas = RegistroFalta.objects.filter(data__month=mes, data__year=ano).select_related('professor').order_by('-data')

    if not faltas.exists():
        messages.warning(request, 'Nenhum registro para exportar neste mês.')
        return redirect('controle_faltas')

    # Cria o arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Faltas {mes}-{ano}"

    # Cabeçalho
    headers = ['Professor', 'Data', 'Dia', 'Previsto', 'Real', 'Status', 'Minutos', 'Observação']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Dados
    for row_num, falta in enumerate(faltas, 2):
        ws.cell(row=row_num, column=1, value=falta.professor.get_full_name() or falta.professor.username)
        ws.cell(row=row_num, column=2, value=falta.data.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=3, value=falta.dia_semana)
        ws.cell(row=row_num, column=4, value=falta.horario_previsto.strftime('%H:%M') if falta.horario_previsto else '-')
        ws.cell(row=row_num, column=5, value=falta.horario_real.strftime('%H:%M') if falta.horario_real else '-')
        ws.cell(row=row_num, column=6, value=falta.get_tipo_display())
        ws.cell(row=row_num, column=7, value=f"{falta.minutos_faltados()} min")
        ws.cell(row=row_num, column=8, value=falta.observacao or '-')

    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = (max_length + 2)

    # Configura a resposta HTTP para download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=faltas_{mes}-{ano}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def listar_relatorios_mensais(request):
    pasta = '/home/gestaoluizzanchim/colegio/media/arquivos_mensais/'
    arquivos = []
    try:
        for arquivo in os.listdir(pasta):
            if arquivo.endswith('.xlsx'):
                data_mod = os.path.getmtime(os.path.join(pasta, arquivo))
                arquivos.append({
                    'nome': arquivo,
                    'url': f'/media/arquivos_mensais/{arquivo}',
                    'data': data_mod,
                })
        arquivos.sort(key=lambda x: x['data'], reverse=True)
    except FileNotFoundError:
        pass
    return render(request, 'lista_relatorios.html', {'arquivos': arquivos})

# =============================================================================
# NOVAS VIEWS PARA TELAS CUSTOMIZADAS
# =============================================================================

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def lista_recados_internos(request):
    if request.method == 'POST':
        form = RecadoInternoForm(request.POST, request.FILES)
        if form.is_valid():
            recado = form.save(commit=False)
            recado.criado_por = request.user
            recado.save()
            return redirect('lista_recados_internos')
    else:
        form = RecadoInternoForm()

    recados = RecadoInterno.objects.all().order_by('-criado_em')
    return render(request, 'lista_recados_internos.html', {
        'recados': recados,
        'form': form
    })

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def lista_documentos_privados(request):
    if request.method == 'POST':
        form = DocumentoPrivadoForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.criado_por = request.user
            doc.save()
            return redirect('lista_documentos_privados')
    else:
        form = DocumentoPrivadoForm()

    documentos = DocumentoPrivado.objects.all().order_by('-criado_em')
    return render(request, 'lista_documentos_privados.html', {
        'documentos': documentos,
        'form': form
    })

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def lista_eventos_privados(request):
    if request.method == 'POST':
        form = EventoPrivadoForm(request.POST)
        if form.is_valid():
            evento = form.save(commit=False)
            evento.criado_por = request.user
            evento.save()
            return redirect('lista_eventos_privados')
    else:
        form = EventoPrivadoForm()

    eventos = EventoPrivado.objects.all().order_by('-data_inicio')
    return render(request, 'lista_eventos_privados.html', {
        'eventos': eventos,
        'form': form
    })

# =============================================================================
# CONTROLE DE FALTAS - ALUNOS
# =============================================================================
from .models import Turma, Aluno, RegistroFaltaAluno

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def controle_faltas_alunos_antigo(request):
    turmas = Turma.objects.filter(ativa=True).order_by('nome')
    mes = request.GET.get('mes', timezone.now().month)
    ano = request.GET.get('ano', timezone.now().year)

    faltas = RegistroFaltaAluno.objects.filter(
        data__month=mes,
        data__year=ano
    ).select_related('aluno', 'aluno__turma').order_by('-data')

    context = {
        'turmas': turmas,
        'faltas': faltas,
        'mes': mes,
        'ano': ano,
        'meses': range(1, 13),
        'anos': [2026, 2027, 2028],
    }
    return render(request, 'controle_faltas_alunos.html', context)

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def registrar_falta_aluno(request):
    if request.method == 'POST':
        aluno_id = request.POST.get('aluno')
        data = request.POST.get('data')
        quantidade = request.POST.get('quantidade', 1)
        justificada = request.POST.get('justificada') == 'on'
        responsavel = request.POST.get('responsavel', '')
        observacoes = request.POST.get('observacoes', '')

        try:
            aluno = Aluno.objects.get(id=aluno_id)
            RegistroFaltaAluno.objects.create(
                aluno=aluno,
                data=data,
                quantidade_faltas=quantidade,
                justificada=justificada,
                responsavel_contatado=responsavel,
                observacoes=observacoes,
                registrado_por=request.user
            )
            messages.success(request, f'Falta registrada para {aluno.nome}')
        except Exception as e:
            messages.error(request, f'Erro: {str(e)}')

        return redirect('controle_faltas_alunos_antigo')

    # GET - exibe formulário
    alunos = Aluno.objects.filter(ativo=True).select_related('turma')
    return render(request, 'registrar_falta_aluno.html', {'alunos': alunos})

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def editar_falta_aluno(request, falta_id):
    falta = get_object_or_404(RegistroFaltaAluno, id=falta_id)

    if request.method == 'POST':
        falta.quantidade_faltas = request.POST.get('quantidade', 1)
        falta.justificada = request.POST.get('justificada') == 'on'
        falta.responsavel_contatado = request.POST.get('responsavel', '')
        falta.observacoes = request.POST.get('observacoes', '')
        falta.save()

        messages.success(request, 'Registro atualizado!')
        return redirect('controle_faltas_alunos_antigo')

    return render(request, 'editar_falta_aluno.html', {'falta': falta})

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def excluir_falta_aluno(request, falta_id):
    falta = get_object_or_404(RegistroFaltaAluno, id=falta_id)
    falta.delete()
    messages.success(request, 'Registro excluído!')
    return redirect('controle_faltas_alunos_antigo')

# =============================================================================
# IMPORTAÇÃO DE ALUNOS VIA EXCEL
# =============================================================================
from openpyxl import load_workbook
from .models import Turma, Aluno

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def importar_alunos_excel(request):
    if request.method == 'POST' and request.FILES.get('arquivo'):
        arquivo = request.FILES['arquivo']
        wb = load_workbook(arquivo)
        ws = wb.active

        contador = 0
        erros = 0

        for row in ws.iter_rows(min_row=2, values_only=True):  # Pula cabeçalho
            try:
                turma_nome = str(row[0]).strip() if row[0] else ''  # Coluna A: Turma (ex: 3A)
                numero = int(row[1]) if row[1] else 0               # Coluna B: Número
                nome = str(row[2]).strip() if row[2] else ''        # Coluna C: Nome do aluno

                if turma_nome and numero and nome:
                    # Extrai a série do nome da turma (ex: 3A -> 3º Ano)
                    serie = f"{turma_nome[0]}º Ano"

                    # Busca ou cria a turma
                    turma, created = Turma.objects.get_or_create(
                        nome=turma_nome,
                        defaults={
                            'serie': serie,
                            'ano': 2026,
                            'ativa': True
                        }
                    )

                    # Cria o aluno
                    Aluno.objects.create(
                        nome=nome,
                        numero=numero,
                        turma=turma
                    )
                    contador += 1
                else:
                    erros += 1
            except Exception as e:
                erros += 1
                print(f"Erro na linha: {e}")

        messages.success(request, f'{contador} alunos importados com sucesso!')
        if erros > 0:
            messages.warning(request, f'{erros} linhas com erro foram ignoradas.')

        return redirect('controle_faltas_alunos')

    return render(request, 'importar_alunos.html')


from django.utils import timezone
from datetime import datetime
from calendar import monthrange

def pertence_ao_grupo_equipe_ou_digitadores(user):
    if not user.is_authenticated:
        return False
    return user.groups.filter(name__in=['Equipe Diretiva', 'Digitadores']).exists()

@login_required
@user_passes_test(pertence_ao_grupo_equipe_ou_digitadores, login_url='/')  # ← CORRETO (novo)
def registrar_ocorrencia_aluno(request):
    # Recupera última data da sessão
    ultima_data_str = request.session.get('ultima_data_ocorrencia')
    if ultima_data_str:
        try:
            ultima_data = datetime.strptime(ultima_data_str, '%Y-%m-%d').date()
        except ValueError:
            ultima_data = timezone.now().date()
    else:
        ultima_data = timezone.now().date()

    # Recupera última turma da sessão
    ultima_turma_id = request.session.get('ultima_turma_ocorrencia_id')
    ultima_turma = None
    if ultima_turma_id:
        try:
            ultima_turma = Turma.objects.get(id=ultima_turma_id)
        except Turma.DoesNotExist:
            ultima_turma = None

    if request.method == 'POST':
        form = RegistroOcorrenciaForm(request.POST)
        if form.is_valid():
            turma = form.cleaned_data['turma']
            numero = form.cleaned_data['numero_aluno']

            aluno = Aluno.objects.filter(turma=turma, numero=numero).first()
            if not aluno:
                messages.error(request, f'Aluno número {numero} não encontrado na turma {turma.nome}!')
                return render(request, 'registrar_ocorrencia.html', {
                    'form': form,
                    'ultimas_ocorrencias': RegistroOcorrenciaAluno.objects.select_related('aluno', 'aluno__turma').order_by('-data', '-horario_chegada')[:10]
                })

            ocorrencia = form.save(commit=False)
            tipo_ocorrencia = request.POST.get('tipo_ocorrencia', 'falta')
            ocorrencia.tipo_ocorrencia = tipo_ocorrencia
            ocorrencia.observacoes_adicionais = request.POST.get('observacoes_adicionais', '')
            ocorrencia.turno = request.POST.get('turno', 'manha')  # ← NOVO CAMPO TURNO
            ocorrencia.faltou = (tipo_ocorrencia == 'falta')  # ← LINHA CRÍTICA

            if ocorrencia.horario_chegada == '':
                ocorrencia.horario_chegada = None
            if ocorrencia.horario_contato == '':
                ocorrencia.horario_contato = None

            ocorrencia.aluno = aluno
            ocorrencia.registrado_por = request.user
            ocorrencia.save()

            # ===== ENVIAR WHATSAPP SE FOR FALTA =====
            if tipo_ocorrencia == 'falta' and aluno.telefone:
                try:
                    mensagem = f"""🏫 *CCM LUIZ ZANCHIM*

⚠️ *OCORRÊNCIA REGISTRADA*

*Aluno:* {aluno.nome}
*Turma:* {aluno.turma.nome}
*Data:* {ocorrencia.data.strftime('%d/%m/%Y')}
*Tipo:* {ocorrencia.get_tipo_ocorrencia_display()}

Para mais detalhes, entre em contato com a Equipe Pedagógica."""

                    enviar_whatsapp(aluno.telefone, mensagem)
                    print(f"📱 WhatsApp enviado para {aluno.nome}")
                except Exception as e:
                    print(f"❌ Erro ao enviar WhatsApp: {e}")

            # ===== SALVA DATA, TURMA, TIPO E TURNO NA SESSÃO =====
            request.session['ultima_data_ocorrencia'] = ocorrencia.data.isoformat()
            request.session['ultima_turma_ocorrencia_id'] = turma.id
            request.session['ultima_turma_ocorrencia_nome'] = turma.nome
            request.session['ultimo_tipo_ocorrencia'] = tipo_ocorrencia
            request.session['ultimo_turno'] = request.POST.get('turno', 'manha')  # ← NOVO

            messages.success(request, f'Ocorrência registrada para {aluno.nome} (Turma {turma.nome}, Nº {numero})')
            return redirect('registrar_ocorrencia_aluno')
        else:
            messages.error(request, 'Erro no formulário. Verifique os dados.')
    else:
        # ===== RECUPERA O ÚLTIMO TIPO E TURNO DA SESSÃO =====
        ultimo_tipo = request.session.get('ultimo_tipo_ocorrencia', 'falta')
        ultimo_turno = request.session.get('ultimo_turno', 'manha')  # ← NOVO

        initial_data = {
            'data': ultima_data.isoformat(),
            'faltou': (ultimo_tipo == 'falta'),
            'tipo_ocorrencia': ultimo_tipo,
            'turno': ultimo_turno,  # ← NOVO
        }
        form = RegistroOcorrenciaForm(initial=initial_data)
        if ultima_turma:
            form.fields['turma'].initial = ultima_turma

    ultimas_ocorrencias = RegistroOcorrenciaAluno.objects.select_related('aluno', 'aluno__turma').order_by('-data', '-horario_chegada')[:10]

    return render(request, 'registrar_ocorrencia.html', {
        'form': form,
        'ultimas_ocorrencias': ultimas_ocorrencias
    })

def buscar_aluno_ajax(request):
    turma_id = request.GET.get('turma_id')
    numero = request.GET.get('numero', '')

    if turma_id and numero:
        try:
            turma = Turma.objects.get(id=turma_id)
            aluno = Aluno.objects.get(turma=turma, numero=numero)
            return JsonResponse({
                'encontrado': True,
                'nome': aluno.nome,
                'turma': turma.nome,
                'numero': aluno.numero
            })
        except (Turma.DoesNotExist, Aluno.DoesNotExist):
            return JsonResponse({
                'encontrado': False,
                'erro': 'Aluno não encontrado nesta turma.'
            })
    return JsonResponse({'encontrado': False, 'erro': 'Informe turma e número.'})

# =============================================================================
# BUSCA ATIVA - GESTÃO DE FALTAS PENDENTES
# =============================================================================

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def busca_ativa(request):
    # Filtros
    mes = int(request.GET.get('mes', timezone.now().month))
    ano = int(request.GET.get('ano', timezone.now().year))
    turma_id = request.GET.get('turma')

    # Busca TODAS as ocorrências (sem filtro faltou=True)
    ocorrencias = RegistroOcorrenciaAluno.objects.filter(
        data__year=ano,
        data__month=mes
    ).select_related('aluno', 'aluno__turma').order_by('-data')

    # Filtra por turma se selecionada
    if turma_id:
        ocorrencias = ocorrencias.filter(aluno__turma_id=turma_id)

    # Lista de turmas para o dropdown
    turmas = Turma.objects.filter(ativa=True).order_by('nome')

    # Resumo por turma
    turmas_resumo = {}
    for occ in ocorrencias:
        turma_nome = occ.aluno.turma.nome
        if turma_nome not in turmas_resumo:
            turmas_resumo[turma_nome] = {'total': 0, 'realizadas': 0}
        turmas_resumo[turma_nome]['total'] += 1
        if occ.busca_ativa_realizada:
            turmas_resumo[turma_nome]['realizadas'] += 1

    context = {
        'ocorrencias': ocorrencias,
        'turmas': turmas,
        'turma_selecionada': turma_id,
        'turmas_resumo': turmas_resumo,
        'mes': mes,
        'ano': ano,
        'meses': range(1, 13),
        'anos': range(2024, 2028),
    }
    return render(request, 'busca_ativa.html', context)


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def marcar_busca_ativa(request, pk):
    """Marca uma falta como 'busca ativa realizada'"""
    ocorrencia = get_object_or_404(RegistroOcorrenciaAluno, id=pk)
    ocorrencia.busca_ativa_realizada = True
    ocorrencia.save()
    messages.success(request, f'✅ Busca ativa marcada para {ocorrencia.aluno.nome}')

    # Pegar os parâmetros da URL para manter os filtros
    mes = request.GET.get('mes', '')
    ano = request.GET.get('ano', '')
    turma_id = request.GET.get('turma', '')

    # Montar a URL de volta com os mesmos filtros
    url = '/busca-ativa/'
    params = []
    if mes:
        params.append(f'mes={mes}')
    if ano:
        params.append(f'ano={ano}')
    if turma_id:
        params.append(f'turma={turma_id}')

    if params:
        url += '?' + '&'.join(params)

    return redirect(url)


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def marcar_todos_busca_ativa(request):
    if request.method == 'POST':
        mes = request.POST.get('mes')
        ano = request.POST.get('ano')
        turma_id = request.POST.get('turma')

        # Remove o filtro faltou=True
        ocorrencias = RegistroOcorrenciaAluno.objects.filter(
            data__year=ano,
            data__month=mes
        )
        if turma_id and turma_id != 'None':
            ocorrencias = ocorrencias.filter(aluno__turma_id=turma_id)

        quantidade = ocorrencias.update(busca_ativa_realizada=True)
        messages.success(request, f'✅ {quantidade} ocorrência(s) marcada(s) como busca ativa realizada!')

    return redirect('busca_ativa')

# =============================================================================
# RELATÓRIO DE FALTAS DE ALUNOS EM EXCEL (DOWNLOAD DIRETO)
# =============================================================================
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from datetime import datetime
from .models import RegistroFaltaAluno

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def relatorio_faltas_alunos(request):
    # Pega os parâmetros da URL (mês e ano)
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')

    # Se não vierem, usa o mês/ano atual
    if not mes or not ano:
        hoje = datetime.now()
        mes = hoje.month
        ano = hoje.year

    # Busca as faltas do período, ordenadas por turma, número e data
    faltas = RegistroFaltaAluno.objects.filter(
        data__month=mes,
        data__year=ano
    ).select_related('aluno', 'aluno__turma').order_by('aluno__turma__nome', 'aluno__numero', 'data')

    # Se não houver faltas, exibe mensagem e volta
    if not faltas.exists():
        messages.warning(request, 'Nenhuma falta encontrada para o período selecionado.')
        return redirect('painel_equipe')

    # Cria o arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Faltas {mes}-{ano}"

    # Cabeçalho
    headers = ['Turma', 'Nº', 'Aluno', 'Data', 'Faltas', 'Justificada', 'Responsável', 'Observações']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Preenche os dados
    for row_num, falta in enumerate(faltas, 2):
        ws.cell(row=row_num, column=1, value=falta.aluno.turma.nome)
        ws.cell(row=row_num, column=2, value=falta.aluno.numero)
        ws.cell(row=row_num, column=3, value=falta.aluno.nome)
        ws.cell(row=row_num, column=4, value=falta.data.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=5, value=falta.quantidade_faltas)
        ws.cell(row=row_num, column=6, value='Sim' if falta.justificada else 'Não')
        ws.cell(row=row_num, column=7, value=falta.responsavel_contatado or '-')
        ws.cell(row=row_num, column=8, value=falta.observacoes or '-')

    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = (max_length + 2)

    # Prepara a resposta HTTP para download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=faltas_alunos_{mes}_{ano}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response

# =============================================================================
# LIMPEZA DO SISTEMA
# =============================================================================

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def limpeza_sistema(request):
    """Executa o script de limpeza do sistema (apenas superusuário)"""
    if not request.user.is_superuser:
        messages.error(request, "Acesso negado. Apenas superusuário.")
        return redirect('painel_equipe')

    try:
        import subprocess
        resultado = subprocess.run(
            ['/home/gestaoluizzanchim/limpeza.sh'],
            capture_output=True,
            text=True,
            timeout=30
        )
        output = resultado.stdout + resultado.stderr

        # Se o script não existir, criar um básico
        if "No such file" in output:
            output = "⚠️ Script de limpeza não encontrado. Execute o comando manualmente:\n\n./limpeza.sh"

        return render(request, 'limpeza_resultado.html', {'output': output})

    except Exception as e:
        messages.error(request, f"Erro ao executar limpeza: {str(e)}")
        return redirect('painel_equipe')

# =============================================================================
# EXPORTAR BUSCA ATIVA PARA EXCEL
# =============================================================================

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def exportar_busca_ativa(request):
    mes = int(request.GET.get('mes', timezone.now().month))
    ano = int(request.GET.get('ano', timezone.now().year))
    turma_id = request.GET.get('turma')

    # 🔴 CORREÇÃO: trata o "None" como string
    if turma_id == 'None' or turma_id == '':
        turma_id = None

    ocorrencias = RegistroOcorrenciaAluno.objects.filter(
        faltou=True,
        data__year=ano,
        data__month=mes
    ).select_related('aluno', 'aluno__turma')

    # Aplica o filtro APENAS se turma_id for um número válido
    if turma_id and turma_id != 'None':
        try:
            turma_id_int = int(turma_id)
            ocorrencias = ocorrencias.filter(aluno__turma_id=turma_id_int)
        except (ValueError, TypeError):
            pass  # Se não for número, ignora o filtro

    # Prepara dados
    data = []
    for occ in ocorrencias:
        data.append([
            occ.data.strftime('%d/%m/%Y'),
            occ.aluno.turma.nome,
            occ.aluno.numero,
            occ.aluno.nome,
            'Sim' if occ.busca_ativa_realizada else 'Não',
            occ.atendido_por or '',
            occ.motivo_alegado or '',
            occ.responsavel_contatado or '',
        ])

    # Se não houver dados, cria uma lista vazia
    if not data:
        data = [['Sem dados para o período', '', '', '', '', '', '', '']]

    df = pd.DataFrame(data, columns=[
        'Data', 'Turma', 'Nº', 'Aluno',
        'Busca Ativa Realizada', 'Atendido por',
        'Motivo Alegado', 'Responsável Contatado'
    ])

    output = BytesIO()
    sheet_name = f'Busca_Ativa_{mes:02d}_{ano}'
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="busca_ativa_{mes:02d}_{ano}.xlsx"'
    return response

# =============================================================================
# RELATÓRIO DE OCORRÊNCIAS DE ALUNOS EM EXCEL
# =============================================================================
@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def relatorio_ocorrencias_alunos(request):
    # Pega os parâmetros da URL (mês e ano)
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')

    # Se não vierem, usa o mês/ano atual
    if not mes or not ano:
        hoje = datetime.now()
        mes = hoje.month
        ano = hoje.year

    # Busca as ocorrências do período
    ocorrencias = RegistroOcorrenciaAluno.objects.filter(
        data__month=mes,
        data__year=ano
    ).select_related('aluno', 'aluno__turma').order_by('-data', '-horario_chegada')

    if not ocorrencias.exists():
        messages.warning(request, 'Nenhuma ocorrência encontrada para o período selecionado.')
        return redirect('painel_equipe')

    # Cria o arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Ocorrências {mes}-{ano}"

    # Cabeçalho - COM COLUNA TIPO
    headers = [
        'Data', 'Turma', 'Nº', 'Aluno', 'Tipo',
        'Horário Chegada', 'Atendido por', 'Motivo (aluno)',
        'Responsável contatado', 'Hora contato', 'Alegado (responsável)'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Preenche os dados
    for row_num, occ in enumerate(ocorrencias, 2):
        ws.cell(row=row_num, column=1, value=occ.data.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=2, value=occ.aluno.turma.nome)
        ws.cell(row=row_num, column=3, value=occ.aluno.numero)
        ws.cell(row=row_num, column=4, value=occ.aluno.nome)
        # COLUNA TIPO (nova)
        ws.cell(row=row_num, column=5, value=occ.get_tipo_ocorrencia_display())
        # Horário Chegada (sem "FALTA")
        ws.cell(row=row_num, column=6, value=occ.horario_chegada.strftime('%H:%M') if occ.horario_chegada else '')
        ws.cell(row=row_num, column=7, value=occ.atendido_por or '')
        ws.cell(row=row_num, column=8, value=occ.motivo_alegado or '')
        ws.cell(row=row_num, column=9, value=occ.responsavel_contatado or '')
        ws.cell(row=row_num, column=10, value=occ.horario_contato.strftime('%H:%M') if occ.horario_contato else '')
        ws.cell(row=row_num, column=11, value=occ.alegado_responsavel or '')

    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ocorrencias_{mes}_{ano}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response

# =============================================================================
# VIEWS PARA CADASTRO E EDIÇÃO DE PROFESSORES E ALUNOS
# =============================================================================

from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User, Group
from .forms import ProfessorForm, ProfessorEditForm, AlunoForm, AlunoEditForm
from .models import Professor, Aluno

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def cadastrar_professor(request):
    if request.method == 'POST':
        form = ProfessorForm(request.POST)
        if form.is_valid():
            # Cria o usuário
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            usuario = User.objects.create_user(
                username=username,
                password=password,
                is_staff=True
            )
            usuario.save()

            # Adiciona ao grupo Professores
            grupo = Group.objects.get(name='Professores')
            usuario.groups.add(grupo)

            # Cria o professor
            professor = form.save(commit=False)
            professor.usuario = usuario
            professor.save()

            messages.success(request, f'✅ Professor {professor.nome_completo} cadastrado com sucesso!')
            return redirect('painel_equipe')
    else:
        form = ProfessorForm()

    return render(request, 'cadastrar_professor.html', {'form': form})


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def cadastrar_aluno(request):
    if request.method == 'POST':
        form = AlunoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '✅ Aluno cadastrado com sucesso!')
            return redirect('painel_equipe')
    else:
        form = AlunoForm()

    return render(request, 'cadastrar_aluno.html', {'form': form})


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def listar_professores(request):
    professores = Professor.objects.all().order_by('nome_completo')
    return render(request, 'listar_professores.html', {'professores': professores})


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def editar_professor(request, professor_id):
    professor = get_object_or_404(Professor, id=professor_id)
    if request.method == 'POST':
        form = ProfessorEditForm(request.POST, instance=professor)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Professor {professor.nome_completo} atualizado!')
            return redirect('listar_professores')
    else:
        form = ProfessorEditForm(instance=professor)

    return render(request, 'editar_professor.html', {'form': form, 'professor': professor})


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def listar_alunos(request):
    alunos = Aluno.objects.all().order_by('turma', 'numero')
    return render(request, 'listar_alunos.html', {'alunos': alunos})


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def editar_aluno(request, aluno_id):
    aluno = get_object_or_404(Aluno, id=aluno_id)
    if request.method == 'POST':
        form = AlunoEditForm(request.POST, instance=aluno)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Aluno {aluno.nome} atualizado!')
            return redirect('listar_alunos')
    else:
        form = AlunoEditForm(instance=aluno)

    return render(request, 'editar_aluno.html', {'form': form, 'aluno': aluno})

# =============================================================================
# RELATÓRIO DE OCORRÊNCIAS POR TIPO (NOVO)
# =============================================================================

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def relatorio_filtro(request):
    from datetime import datetime
    context = {
        'meses': range(1, 13),
        'anos': range(2024, 2028),
        'mes_atual': datetime.now().month,
        'ano_atual': datetime.now().year,
    }
    return render(request, 'relatorio_filtro.html', context)


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def relatorio_ocorrencias_por_tipo(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime
    from django.http import HttpResponse

    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    tipo = request.GET.get('tipo', 'todas')

    if not mes or not ano:
        hoje = datetime.now()
        mes = hoje.month
        ano = hoje.year
    else:
        mes = int(mes)
        ano = int(ano)

    tipo_nomes = {
        'todas': 'Todas as ocorrências',
        'falta': 'Falta',
        'atraso': 'Atraso',
        'piercing': 'Uso de Piercing',
        'cabelo': 'Cabelo',
        'uniforme': 'Uniforme',
        'desvio_normas': 'Desvio de Normas',
    }

    ocorrencias = RegistroOcorrenciaAluno.objects.filter(
        data__month=mes,
        data__year=ano
    ).select_related('aluno', 'aluno__turma').order_by('-data', '-horario_chegada')

    if tipo != 'todas':
        ocorrencias = ocorrencias.filter(tipo_ocorrencia=tipo)

    if not ocorrencias.exists():
        messages.warning(request, f'Nenhuma ocorrência do tipo "{tipo_nomes.get(tipo, tipo)}" encontrada para o período.')
        return redirect('painel_equipe')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{tipo_nomes.get(tipo, tipo)}_{mes}_{ano}"

    # HEADERS COM COLUNA TIPO
    if tipo == 'desvio_normas':
        headers = ['Data', 'Turma', 'Nº', 'Aluno', 'Tipo', 'Atendido por',
                   'Motivo', 'Responsável', 'Alegado', 'ATA']
    else:
        headers = ['Data', 'Turma', 'Nº', 'Aluno', 'Tipo', 'Atendido por',
                   'Motivo', 'Responsável', 'Alegado']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    for row_num, occ in enumerate(ocorrencias, 2):
        ws.cell(row=row_num, column=1, value=occ.data.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=2, value=occ.aluno.turma.nome)
        ws.cell(row=row_num, column=3, value=occ.aluno.numero)
        ws.cell(row=row_num, column=4, value=occ.aluno.nome)
        ws.cell(row=row_num, column=5, value=occ.get_tipo_ocorrencia_display())  # TIPO
        ws.cell(row=row_num, column=6, value=occ.atendido_por or '')
        ws.cell(row=row_num, column=7, value=occ.motivo_alegado or '')
        ws.cell(row=row_num, column=8, value=occ.responsavel_contatado or '')
        ws.cell(row=row_num, column=9, value=occ.alegado_responsavel or '')

        if tipo == 'desvio_normas':
            ata_texto = occ.observacoes_adicionais or ''
            if len(ata_texto) > 200:
                ata_texto = ata_texto[:200] + '...'
            ws.cell(row=row_num, column=10, value=ata_texto)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nome_arquivo = f"relatorio_{tipo}_{mes}_{ano}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={nome_arquivo}'

    wb.save(response)
    return response


# =============================================================================
# CADASTRO E EDIÇÃO DE ALUNOS (já existente)
# =============================================================================

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def editar_aluno(request, aluno_id):
    aluno = get_object_or_404(Aluno, id=aluno_id)
    if request.method == 'POST':
        form = AlunoEditForm(request.POST, instance=aluno)
        if form.is_valid():
            form.save()
            messages.success(request, f'✅ Aluno {aluno.nome} atualizado!')
            return redirect('listar_alunos')
    else:
        form = AlunoEditForm(instance=aluno)

    return render(request, 'editar_aluno.html', {'form': form, 'aluno': aluno})

# =============================================================================
# GESTÃO DE LABORATÓRIOS E EMPRÉSTIMOS
# =============================================================================

from .models import Laboratorio, ItemEquipamento, AgendamentoLab, Emprestimo
from django.db.models import Q
from datetime import datetime, timedelta

@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def listar_laboratorios(request):
    """Lista todos os laboratórios"""
    laboratorios = Laboratorio.objects.filter(ativo=True)

    # Contagem de disponíveis para itinerantes
    for lab in laboratorios:
        if lab.tipo == 'itinerante':
            lab.disponiveis = ItemEquipamento.objects.filter(
                laboratorio=lab,
                disponivel=True
            ).count()

    return render(request, 'laboratorios/listar.html', {'laboratorios': laboratorios})


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def agendamento_lab(request, lab_id):
    """Agendar laboratório fixo (Lab 01, 02, 03)"""
    laboratorio = get_object_or_404(Laboratorio, id=lab_id, tipo='fixo')

    if request.method == 'POST':
        data = request.POST.get('data')
        horario = request.POST.get('horario')
        turno = request.POST.get('turno')
        professor_id = request.POST.get('professor')
        disciplina = request.POST.get('disciplina')
        turma_id = request.POST.get('turma')
        observacao = request.POST.get('observacao', '')

        # 1. Verificar se o mesmo professor já tem agendamento neste horário (em qualquer lab)
        conflito_professor = AgendamentoLab.objects.filter(
            professor_id=professor_id,
            data=data,
            horario=horario,
            turno=turno
        ).exists()

        if conflito_professor:
            messages.error(request, '❌ Este professor já possui agendamento neste horário em outro laboratório!')
            # Buscar dados para renderizar o formulário novamente
            from .models import Professor, Turma
            professores = Professor.objects.filter(ativo=True)
            turmas = Turma.objects.filter(ativa=True)
            return render(request, 'laboratorios/agendar.html', {
                'laboratorio': laboratorio,
                'professores': professores,
                'turmas': turmas,
            })

        # 2. Verificar se o laboratório já está reservado neste horário
        conflito_lab = AgendamentoLab.objects.filter(
            laboratorio=laboratorio,
            data=data,
            horario=horario,
            turno=turno
        ).exists()

        if conflito_lab:
            messages.error(request, f'❌ {laboratorio.nome} já está reservado neste horário!')
            from .models import Professor, Turma
            professores = Professor.objects.filter(ativo=True)
            turmas = Turma.objects.filter(ativa=True)
            return render(request, 'laboratorios/agendar.html', {
                'laboratorio': laboratorio,
                'professores': professores,
                'turmas': turmas,
            })

        # 3. Salvar agendamento
        from .models import Professor, Turma
        AgendamentoLab.objects.create(
            laboratorio=laboratorio,
            data=data,
            horario=horario,
            turno=turno,
            professor_id=professor_id,
            disciplina=disciplina,
            turma_id=turma_id,
            observacao=observacao,
            registrado_por=request.user
        )
        messages.success(request, f'✅ {laboratorio.nome} agendado com sucesso!')
        return redirect('listar_laboratorios')

    # GET - mostrar formulário
    from .models import Professor, Turma
    professores = Professor.objects.filter(ativo=True)
    turmas = Turma.objects.filter(ativa=True)

    return render(request, 'laboratorios/agendar.html', {
        'laboratorio': laboratorio,
        'professores': professores,
        'turmas': turmas,
    })


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def cronograma_semanal(request):
    """Exibe o cronograma semanal dos laboratórios fixos"""
    from datetime import datetime, timedelta

    # Pegar a semana (padrão: semana atual)
    data_inicio = request.GET.get('data_inicio')
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    else:
        data_inicio = datetime.now().date()
        data_inicio = data_inicio - timedelta(days=data_inicio.weekday())

    data_fim = data_inicio + timedelta(days=6)

    # Buscar agendamentos da semana
    agendamentos = AgendamentoLab.objects.filter(
        data__gte=data_inicio,
        data__lte=data_fim
    ).select_related('laboratorio', 'professor', 'turma').order_by('data', 'horario')

    # Mapa de dias em português
    mapa_dias = {
        'Monday': 'Segunda-feira',
        'Tuesday': 'Terça-feira',
        'Wednesday': 'Quarta-feira',
        'Thursday': 'Quinta-feira',
        'Friday': 'Sexta-feira'
    }

    # Criar dicionário para acesso rápido
    agendamentos_dict = {}
    for a in agendamentos:
        dia_en = a.data.strftime('%A')
        dia_pt = mapa_dias.get(dia_en, dia_en)
        key = (a.laboratorio.id, dia_pt, a.horario)
        agendamentos_dict[key] = a

    dias_semana = ['Segunda-feira', 'Terça-feira', 'Quarta-feira', 'Quinta-feira', 'Sexta-feira']
    horarios = ['1', '2', '3', '4', '5', '6']

    context = {
        'dias_semana': dias_semana,
        'horarios': horarios,
        'agendamentos_dict': agendamentos_dict,
        'laboratorios': Laboratorio.objects.filter(tipo='fixo', ativo=True),
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'semana_anterior': data_inicio - timedelta(days=7),
        'semana_proxima': data_inicio + timedelta(days=7),
    }
    return render(request, 'laboratorios/cronograma.html', context)


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def emprestimo_equipamento(request, lab_id):
    """Registrar empréstimo de equipamentos (Lab 04 ou Lab 05)"""
    laboratorio = get_object_or_404(Laboratorio, id=lab_id, tipo='itinerante')

    # Equipamentos disponíveis
    equipamentos_disponiveis = ItemEquipamento.objects.filter(
        laboratorio=laboratorio,
        disponivel=True
    )

    if request.method == 'POST':
        tipo_emprestimo = request.POST.get('tipo_emprestimo')
        quantidade = int(request.POST.get('quantidade', 0))
        data_prevista = request.POST.get('data_prevista_devolucao')
        motivo = request.POST.get('motivo')
        aulas_utilizacao = request.POST.get('aulas_utilizacao', '')
        observacao = request.POST.get('observacao', '')

        # Pegar equipamentos selecionados (checkbox)
        itens_selecionados = request.POST.getlist('itens')

        # Validar
        if len(itens_selecionados) != quantidade:
            messages.error(request, 'Selecione a quantidade correta de equipamentos!')
        else:
            # Criar empréstimo
            emprestimo = Emprestimo.objects.create(
                tipo_emprestimo=tipo_emprestimo,
                laboratorio=laboratorio,
                quantidade=quantidade,
                data_prevista_devolucao=data_prevista,
                motivo=motivo,
                aulas_utilizacao=aulas_utilizacao,
                observacao=observacao,
                registrado_por=request.user
            )

            # Adicionar itens e marcar como indisponíveis
            for item_id in itens_selecionados:
                item = ItemEquipamento.objects.get(id=item_id)
                emprestimo.itens.add(item)
                item.disponivel = False
                item.save()

            # Se for professor
            if tipo_emprestimo == 'professor':
                emprestimo.professor_id = request.POST.get('professor_id')
            elif tipo_emprestimo == 'aluno':
                emprestimo.aluno_id = request.POST.get('aluno_id')
            elif tipo_emprestimo == 'turma':
                emprestimo.turma_id = request.POST.get('turma_id')

            emprestimo.save()
            messages.success(request, f'Empréstimo registrado! {quantidade} equipamento(s) emprestado(s).')
            return redirect('listar_laboratorios')

    # GET - mostrar formulário
    from .models import Professor, Aluno, Turma
    professores = Professor.objects.filter(ativo=True)
    alunos = Aluno.objects.filter(ativo=True)
    turmas = Turma.objects.filter(ativa=True)

    return render(request, 'laboratorios/emprestimo.html', {
        'laboratorio': laboratorio,
        'equipamentos': equipamentos_disponiveis,
        'professores': professores,
        'alunos': alunos,
        'turmas': turmas,
    })


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def devolver_equipamento(request, emprestimo_id):
    """Registrar devolução de equipamentos"""
    emprestimo = get_object_or_404(Emprestimo, id=emprestimo_id)

    if request.method == 'POST':
        from django.utils import timezone
        emprestimo.data_devolucao = timezone.now()
        emprestimo.status = 'devolvido'
        emprestimo.save()

        # Marcar equipamentos como disponíveis novamente
        for item in emprestimo.itens.all():
            item.disponivel = True
            item.save()

        messages.success(request, 'Equipamentos devolvidos com sucesso!')
        return redirect('listar_emprestimos')

    return render(request, 'laboratorios/devolver.html', {'emprestimo': emprestimo})


@login_required
@user_passes_test(pertence_ao_grupo_equipe_diretiva, login_url='/')
def listar_emprestimos(request):
    """Lista todos os empréstimos ativos"""
    emprestimos = Emprestimo.objects.filter(status='emprestado').order_by('data_prevista_devolucao')
    return render(request, 'laboratorios/lista_emprestimos.html', {'emprestimos': emprestimos})


# ===== WEBHOOK PARA WHATSAPP =====
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse
from django.utils import timezone

@csrf_exempt
def webhook_whatsapp(request):
    """Recebe respostas dos pais via WhatsApp"""
    if request.method == 'POST':
        dados = request.POST
        numero_pai = dados.get('From', '').replace('whatsapp:', '')
        mensagem_texto = dados.get('Body', '')

        from .models import Aluno, RegistroOcorrenciaAluno

        # Buscar aluno pelo telefone
        aluno = Aluno.objects.filter(telefone=numero_pai).first()

        if aluno:
            # Buscar a falta mais recente sem resposta
            ocorrencia = RegistroOcorrenciaAluno.objects.filter(
                aluno=aluno,
                faltou=True,
                resposta_disparo__isnull=True
            ).order_by('-data').first()

            if ocorrencia:
                # Salvar a resposta
                ocorrencia.resposta_disparo = mensagem_texto
                ocorrencia.resposta_disparo_data = timezone.now()
                ocorrencia.save()

                print(f"✅ Resposta WhatsApp salva para {aluno.nome}")

        return HttpResponse('OK', status=200)

    return JsonResponse({'erro': 'Método não permitido'}, status=405)